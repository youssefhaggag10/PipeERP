from datetime import date

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font
from PySide6.QtCore import QDate, Qt
from PySide6.QtWidgets import (
    QComboBox,
    QDateEdit,
    QFileDialog,
    QFormLayout,
    QHBoxLayout,
    QHeaderView,
    QLabel,
    QMessageBox,
    QPushButton,
    QTableWidget,
    QTableWidgetItem,
    QVBoxLayout,
    QWidget,
)

from app.repositories.accounting_repository import AccountingRepository
from app.repositories.partner_repository import PartnerRepository


class ReportsPage(QWidget):
    def __init__(
        self,
        accounting_repository: AccountingRepository,
        partner_repository: PartnerRepository,
    ) -> None:
        super().__init__()
        self.accounting_repository = accounting_repository
        self.partner_repository = partner_repository
        self.current_columns: list[str] = []
        self.current_rows: list[dict] = []
        self.setLayoutDirection(Qt.RightToLeft)

        title = QLabel("التقارير")
        title.setObjectName("titleLabel")
        subtitle = QLabel("تقارير محاسبية وتشغيلية مع فلترة بالعميل أو المورد وتصدير Excel")
        subtitle.setObjectName("subtitleLabel")

        self.report_type = QComboBox()
        self.report_type.addItem("تقرير المبيعات", "sales")
        self.report_type.addItem("تقرير المشتريات", "purchases")
        self.report_type.addItem("أرصدة العملاء", "customer_balances")
        self.report_type.addItem("أرصدة الموردين", "supplier_balances")
        self.report_type.addItem("حركات التحصيل والسداد", "payments")
        self.report_type.addItem("تقييم المخزون", "inventory_valuation")
        self.report_type.currentIndexChanged.connect(self._reload_partner_filter)

        self.partner_filter = QComboBox()
        self.date_from = QDateEdit()
        self.date_from.setCalendarPopup(True)
        self.date_from.setDate(QDate.currentDate().addMonths(-1))
        self.date_to = QDateEdit()
        self.date_to.setCalendarPopup(True)
        self.date_to.setDate(QDate.currentDate())

        filters = QFormLayout()
        filters.addRow("نوع التقرير", self.report_type)
        filters.addRow("العميل / المورد", self.partner_filter)
        filters.addRow("من تاريخ", self.date_from)
        filters.addRow("إلى تاريخ", self.date_to)

        generate_button = QPushButton("إنشاء التقرير")
        generate_button.clicked.connect(self.generate_report)
        export_button = QPushButton("EXPORT XLSX")
        export_button.clicked.connect(self.export_xlsx)
        actions = QHBoxLayout()
        actions.addWidget(generate_button)
        actions.addWidget(export_button)
        actions.addStretch()

        self.summary_label = QLabel("لا توجد بيانات معروضة")
        self.summary_label.setStyleSheet("font-size: 16px; font-weight: 700; color: #38BDF8;")
        self.table = QTableWidget()
        self.table.setEditTriggers(QTableWidget.NoEditTriggers)
        self.table.horizontalHeader().setSectionResizeMode(QHeaderView.Stretch)

        layout = QVBoxLayout(self)
        layout.setContentsMargins(24, 24, 24, 24)
        layout.addWidget(title)
        layout.addWidget(subtitle)
        layout.addLayout(filters)
        layout.addLayout(actions)
        layout.addWidget(self.summary_label)
        layout.addWidget(self.table)
        self._reload_partner_filter()
        self.generate_report()

    def reload(self) -> None:
        self._reload_partner_filter()
        self.generate_report()

    def _reload_partner_filter(self) -> None:
        current = self.partner_filter.currentData()
        report_key = self.report_type.currentData()
        self.partner_filter.clear()
        self.partner_filter.addItem("الكل", None)
        if report_key in {"sales", "customer_balances"}:
            partner_type = "customer"
        elif report_key in {"purchases", "supplier_balances"}:
            partner_type = "supplier"
        elif report_key == "payments":
            partner_type = None
        else:
            self.partner_filter.setEnabled(False)
            return
        self.partner_filter.setEnabled(True)
        partner_rows = []
        if partner_type is None:
            partner_rows.extend(self.partner_repository.list_partners("customer"))
            partner_rows.extend(self.partner_repository.list_partners("supplier"))
        else:
            partner_rows = self.partner_repository.list_partners(partner_type)
        for partner in sorted(partner_rows, key=lambda item: str(item["name"])):
            self.partner_filter.addItem(partner["name"], partner["id"])
        if current is not None:
            index = self.partner_filter.findData(current)
            if index >= 0:
                self.partner_filter.setCurrentIndex(index)

    def generate_report(self) -> None:
        start = date(
            self.date_from.date().year(),
            self.date_from.date().month(),
            self.date_from.date().day(),
        )
        end = date(
            self.date_to.date().year(),
            self.date_to.date().month(),
            self.date_to.date().day(),
        )
        if start > end:
            QMessageBox.warning(self, "تنبيه", "تاريخ البداية يجب ألا يكون بعد تاريخ النهاية")
            return
        try:
            columns, rows = self.accounting_repository.generate_report(
                str(self.report_type.currentData()),
                date_from=start,
                date_to=end,
                partner_id=self.partner_filter.currentData(),
            )
        except ValueError as error:
            QMessageBox.warning(self, "تنبيه", str(error))
            return
        self.current_columns = columns
        self.current_rows = rows
        self.table.setColumnCount(len(columns))
        self.table.setHorizontalHeaderLabels(columns)
        self.table.setRowCount(len(rows))
        for row_index, row in enumerate(rows):
            for column_index, column in enumerate(columns):
                value = row.get(column, "")
                if isinstance(value, float):
                    display = f"{value:,.2f}"
                else:
                    display = str(value)
                self.table.setItem(row_index, column_index, QTableWidgetItem(display))
        self.summary_label.setText(f"عدد السجلات: {len(rows):,}")

    def export_xlsx(self) -> None:
        if not self.current_columns:
            QMessageBox.warning(self, "تنبيه", "أنشئ التقرير أولًا")
            return
        default_name = f"{self.report_type.currentData()}_{QDate.currentDate().toString('yyyyMMdd')}.xlsx"
        path, _ = QFileDialog.getSaveFileName(
            self,
            "حفظ تقرير Excel",
            default_name,
            "Excel Workbook (*.xlsx)",
        )
        if not path:
            return
        if not path.lower().endswith(".xlsx"):
            path += ".xlsx"

        workbook = Workbook()
        sheet = workbook.active
        sheet.title = "التقرير"
        sheet.sheet_view.rightToLeft = True
        for column_index, column in enumerate(self.current_columns, start=1):
            cell = sheet.cell(row=1, column=column_index, value=column)
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal="center")
        for row_index, row in enumerate(self.current_rows, start=2):
            for column_index, column in enumerate(self.current_columns, start=1):
                sheet.cell(row=row_index, column=column_index, value=row.get(column, ""))
        sheet.freeze_panes = "A2"
        sheet.auto_filter.ref = sheet.dimensions
        for column_cells in sheet.columns:
            max_length = max(len(str(cell.value or "")) for cell in column_cells)
            sheet.column_dimensions[column_cells[0].column_letter].width = min(max_length + 4, 45)
        try:
            workbook.save(path)
        except OSError as error:
            QMessageBox.critical(self, "خطأ", f"تعذر حفظ الملف: {error}")
            return
        QMessageBox.information(self, "تم", "تم تصدير التقرير إلى XLSX بنجاح")
