from __future__ import annotations

from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font
from PySide6.QtCore import QDate, Qt
from PySide6.QtWidgets import (
    QComboBox,
    QDateEdit,
    QFileDialog,
    QFormLayout,
    QGroupBox,
    QHBoxLayout,
    QHeaderView,
    QLabel,
    QMessageBox,
    QPushButton,
    QTreeWidget,
    QTreeWidgetItem,
    QVBoxLayout,
    QWidget,
)

from app.repositories.print_settings_repository import PrintSettingsRepository
from app.services.a4_print_service import A4PrintService


class CustomerStatementPage(QWidget):
    """Independent customer ledger screen inside Accounts."""

    def __init__(self, repository, partner_repository, parent=None) -> None:
        super().__init__(parent)
        self.repository = repository
        self.partner_repository = partner_repository
        self.print_settings_repository = PrintSettingsRepository(repository.database)
        self.print_service = A4PrintService()
        self.current_statement: dict | None = None
        self.setLayoutDirection(Qt.RightToLeft)
        self.setObjectName("customerStatementPage")

        filters = QGroupBox("خيارات كشف الحساب")
        form = QFormLayout(filters)
        form.setFieldGrowthPolicy(QFormLayout.FieldGrowthPolicy.AllNonFixedFieldsGrow)
        self.customer_input = QComboBox()
        self.customer_input.setObjectName("statementCustomerInput")
        today = QDate.currentDate()
        self.date_from_input = QDateEdit(QDate(today.year(), today.month(), 1))
        self.date_from_input.setCalendarPopup(True)
        self.date_from_input.setDisplayFormat("yyyy-MM-dd")
        self.date_to_input = QDateEdit(today)
        self.date_to_input.setCalendarPopup(True)
        self.date_to_input.setDisplayFormat("yyyy-MM-dd")
        self.document_status_input = QComboBox()
        self.document_status_input.addItem("المعتمدة فقط", False)
        self.document_status_input.addItem("إظهار المسودات للمراجعة", True)
        self.statement_type_input = QComboBox()
        self.statement_type_input.addItem("مختصر", False)
        self.statement_type_input.addItem("تفصيلي", True)
        form.addRow("العميل", self.customer_input)
        form.addRow("من تاريخ", self.date_from_input)
        form.addRow("إلى تاريخ", self.date_to_input)
        form.addRow("حالة المستندات", self.document_status_input)
        form.addRow("نوع الكشف", self.statement_type_input)

        actions = QHBoxLayout()
        view_button = QPushButton("عرض")
        view_button.clicked.connect(self.show_statement)
        preview_button = QPushButton("معاينة وطباعة")
        preview_button.clicked.connect(self.preview_statement)
        pdf_button = QPushButton("تصدير PDF")
        pdf_button.clicked.connect(self.export_pdf)
        excel_button = QPushButton("تصدير Excel")
        excel_button.clicked.connect(self.export_excel)
        refresh_button = QPushButton("تحديث العملاء")
        refresh_button.setObjectName("secondaryButton")
        refresh_button.clicked.connect(self.reload)
        for button in (view_button, preview_button, pdf_button, excel_button, refresh_button):
            actions.addWidget(button)
        actions.addStretch()

        self.statement_table = QTreeWidget()
        self.statement_table.setObjectName("customerStatementTable")
        self.statement_table.setColumnCount(8)
        self.statement_table.setHeaderLabels(
            [
                "التاريخ",
                "رقم المستند",
                "نوع المستند",
                "البيان",
                "مدين",
                "دائن",
                "الرصيد الجاري",
                "حالة المستند",
            ]
        )
        self.statement_table.setAlternatingRowColors(True)
        self.statement_table.setRootIsDecorated(True)
        self.statement_table.setUniformRowHeights(False)
        self.statement_table.setSelectionMode(QTreeWidget.SelectionMode.SingleSelection)
        self.statement_table.header().setSectionResizeMode(
            QHeaderView.ResizeMode.ResizeToContents
        )
        self.statement_table.header().setSectionResizeMode(
            3, QHeaderView.ResizeMode.Stretch
        )

        summary_group = QGroupBox("ملخص كشف حساب العميل")
        summary_layout = QHBoxLayout(summary_group)
        self.summary_labels: dict[str, QLabel] = {}
        summary_fields = (
            ("opening_balance", "رصيد أول المدة"),
            ("standard_sales_total", "فواتير البيع العادية"),
            ("weight_sales_total", "فواتير البيع بالوزن"),
            ("returns_total", "المرتجعات"),
            ("receipts_total", "التحصيلات"),
            ("adjustments_total", "الخصومات والتسويات"),
            ("net_movement", "صافي الحركة"),
            ("closing_balance", "الرصيد النهائي المستحق"),
        )
        for key, title in summary_fields:
            container = QGroupBox(title)
            value = QLabel("0.00")
            value.setAlignment(Qt.AlignmentFlag.AlignCenter)
            value.setStyleSheet("font-size:16px;font-weight:800;")
            box = QVBoxLayout(container)
            box.addWidget(value)
            summary_layout.addWidget(container)
            self.summary_labels[key] = value

        layout = QVBoxLayout(self)
        layout.addWidget(filters)
        layout.addLayout(actions)
        layout.addWidget(self.statement_table, 1)
        layout.addWidget(summary_group)
        self.reload()

    def reload(self) -> None:
        selected = self.customer_input.currentData()
        self.customer_input.blockSignals(True)
        self.customer_input.clear()
        self.customer_input.addItem("اختر العميل", None)
        for customer in self.partner_repository.list_partners("customer"):
            label = str(customer["name"])
            code = str(customer.get("code") or "").strip()
            self.customer_input.addItem(
                f"{label} — {code}" if code else label,
                int(customer["id"]),
            )
        if selected is not None:
            index = self.customer_input.findData(selected)
            if index >= 0:
                self.customer_input.setCurrentIndex(index)
        self.customer_input.blockSignals(False)

    def _statement_arguments(self) -> dict:
        customer_id = self.customer_input.currentData()
        if customer_id is None:
            raise ValueError("اختر العميل")
        return {
            "customer_id": int(customer_id),
            "date_from": self.date_from_input.date().toPython(),
            "date_to": self.date_to_input.date().toPython(),
            "include_drafts": bool(self.document_status_input.currentData()),
            "detailed": bool(self.statement_type_input.currentData()),
        }

    def show_statement(self) -> None:
        try:
            self.current_statement = self.repository.get_customer_statement(
                **self._statement_arguments()
            )
        except ValueError as error:
            QMessageBox.warning(self, "تعذر عرض الكشف", str(error))
            return
        self._fill_table(self.current_statement)
        self._fill_summary(self.current_statement.get("summary", {}))

    def _ensure_statement(self) -> dict | None:
        try:
            arguments = self._statement_arguments()
        except ValueError as error:
            QMessageBox.warning(self, "تنبيه", str(error))
            return None
        if self.current_statement is None or any(
            (
                self.current_statement.get("customer", {}).get("id")
                != arguments["customer_id"],
                self.current_statement.get("date_from")
                != arguments["date_from"].isoformat(),
                self.current_statement.get("date_to")
                != arguments["date_to"].isoformat(),
                bool(self.current_statement.get("include_drafts"))
                != arguments["include_drafts"],
                bool(self.current_statement.get("detailed"))
                != arguments["detailed"],
            )
        ):
            self.show_statement()
        return self.current_statement

    def _fill_table(self, statement: dict) -> None:
        self.statement_table.clear()
        for movement in statement.get("movements", []):
            item = QTreeWidgetItem(
                [
                    str(movement.get("movement_date", ""))[:10],
                    str(movement.get("document_number", "")),
                    str(movement.get("document_type", "")),
                    str(movement.get("description", "")),
                    self._money(movement.get("debit")),
                    self._money(movement.get("credit")),
                    self._money(movement.get("running_balance")),
                    str(movement.get("status", "")),
                ]
            )
            self.statement_table.addTopLevelItem(item)
            for line in movement.get("lines", []) or []:
                actual_weight = line.get("actual_weight_kg")
                price = line.get("price_per_kg", line.get("unit_price", 0))
                details = [
                    str(line.get("description", "") or ""),
                    f"الكمية {float(line.get('quantity', 0) or 0):g} "
                    f"{line.get('unit', '')}",
                ]
                if actual_weight is not None:
                    details.extend(
                        [
                            f"وزن الكارتة {float(actual_weight):,.3f} كجم",
                            f"سعر الكيلو {float(price or 0):,.2f}",
                        ]
                    )
                else:
                    details.append(f"سعر الوحدة {float(price or 0):,.2f}")
                details.append(
                    f"إجمالي البند {float(line.get('line_total', 0) or 0):,.2f}"
                )
                if line.get("notes"):
                    details.append(str(line["notes"]))
                child = QTreeWidgetItem(
                    [
                        "",
                        "↳",
                        "بند فاتورة",
                        " — ".join(details),
                        "",
                        "",
                        "",
                        "",
                    ]
                )
                item.addChild(child)
            if movement.get("lines"):
                item.setExpanded(True)

    def _fill_summary(self, summary: dict) -> None:
        for key, label in self.summary_labels.items():
            label.setText(self._money(summary.get(key)))

    def preview_statement(self) -> None:
        statement = self._ensure_statement()
        if statement is None:
            return
        try:
            self.print_service.preview_customer_statement(
                statement,
                self.print_settings_repository.get_settings(),
                self,
            )
        except (ValueError, RuntimeError) as error:
            QMessageBox.warning(self, "تعذر الطباعة", str(error))

    def export_pdf(self) -> None:
        statement = self._ensure_statement()
        if statement is None:
            return
        customer = statement.get("customer", {})
        default_name = f"Customer-Statement-{customer.get('code') or customer.get('id')}.pdf"
        path, _ = QFileDialog.getSaveFileName(
            self,
            "تصدير كشف حساب العميل PDF",
            str(Path.home() / default_name),
            "PDF Files (*.pdf)",
        )
        if not path:
            return
        try:
            exported = self.print_service.export_customer_statement_pdf(
                statement,
                self.print_settings_repository.get_settings(),
                path,
            )
        except (ValueError, RuntimeError) as error:
            QMessageBox.warning(self, "تعذر التصدير", str(error))
            return
        QMessageBox.information(self, "تم", f"تم تصدير كشف الحساب إلى:\n{exported}")

    def export_excel(self) -> None:
        statement = self._ensure_statement()
        if statement is None:
            return
        customer = statement.get("customer", {})
        default_name = f"Customer-Statement-{customer.get('code') or customer.get('id')}.xlsx"
        path, _ = QFileDialog.getSaveFileName(
            self,
            "تصدير كشف حساب العميل Excel",
            str(Path.home() / default_name),
            "Excel Files (*.xlsx)",
        )
        if not path:
            return
        try:
            exported = self._write_excel(statement, Path(path))
        except (OSError, ValueError) as error:
            QMessageBox.warning(self, "تعذر التصدير", str(error))
            return
        QMessageBox.information(self, "تم", f"تم تصدير كشف الحساب إلى:\n{exported}")

    @staticmethod
    def _write_excel(statement: dict, path: Path) -> Path:
        output = path.expanduser().resolve()
        if output.suffix.lower() != ".xlsx":
            output = output.with_suffix(".xlsx")
        output.parent.mkdir(parents=True, exist_ok=True)
        workbook = Workbook()
        sheet = workbook.active
        sheet.title = "كشف الحساب"
        sheet.sheet_view.rightToLeft = True
        customer = statement.get("customer", {})
        sheet.append(["كشف حساب عميل"])
        sheet.append(["العميل", customer.get("name", "")])
        sheet.append(["الكود", customer.get("code", "")])
        sheet.append(["من تاريخ", statement.get("date_from", "")])
        sheet.append(["إلى تاريخ", statement.get("date_to", "")])
        sheet.append([])
        headers = [
            "التاريخ",
            "رقم المستند",
            "نوع المستند",
            "البيان",
            "مدين",
            "دائن",
            "الرصيد الجاري",
            "الحالة",
        ]
        sheet.append(headers)
        header_row = sheet.max_row
        for cell in sheet[header_row]:
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal="center")
        for movement in statement.get("movements", []):
            sheet.append(
                [
                    str(movement.get("movement_date", ""))[:10],
                    movement.get("document_number", ""),
                    movement.get("document_type", ""),
                    movement.get("description", ""),
                    float(movement.get("debit", 0) or 0),
                    float(movement.get("credit", 0) or 0),
                    float(movement.get("running_balance", 0) or 0),
                    movement.get("status", ""),
                ]
            )
            for line in movement.get("lines", []) or []:
                actual = line.get("actual_weight_kg")
                price = line.get("price_per_kg", line.get("unit_price", 0))
                description = (
                    f"↳ {line.get('description', '')} — "
                    f"الكمية {float(line.get('quantity', 0) or 0):g} "
                    f"{line.get('unit', '')} — "
                )
                if actual is not None:
                    description += (
                        f"الوزن {float(actual):,.3f} كجم — "
                        f"سعر الكيلو {float(price or 0):,.2f} — "
                    )
                else:
                    description += f"سعر الوحدة {float(price or 0):,.2f} — "
                description += f"الإجمالي {float(line.get('line_total', 0) or 0):,.2f}"
                if line.get("notes"):
                    description += f" — {line['notes']}"
                sheet.append(["", "↳", "بند فاتورة", description, "", "", "", ""])
        sheet.append([])
        sheet.append(["الملخص"])
        sheet[sheet.max_row][0].font = Font(bold=True)
        summary_titles = {
            "opening_balance": "رصيد أول المدة",
            "standard_sales_total": "فواتير البيع العادية",
            "weight_sales_total": "فواتير البيع بالوزن",
            "returns_total": "المرتجعات",
            "receipts_total": "التحصيلات",
            "adjustments_total": "الخصومات والتسويات",
            "net_movement": "صافي الحركة",
            "closing_balance": "الرصيد النهائي المستحق",
        }
        for key, title in summary_titles.items():
            sheet.append([title, float(statement.get("summary", {}).get(key, 0) or 0)])
        widths = [14, 18, 22, 70, 16, 16, 18, 24]
        for index, width in enumerate(widths, start=1):
            sheet.column_dimensions[chr(64 + index)].width = width
        workbook.save(output)
        if not output.exists() or output.stat().st_size <= 0:
            raise ValueError("تعذر إنشاء ملف Excel")
        return output

    @staticmethod
    def _money(value: object) -> str:
        return f"{float(value or 0):,.2f}"


__all__ = ["CustomerStatementPage"]
